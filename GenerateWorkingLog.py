# Filename = Generate Working Log 4 SCT
# Author: dwa

# use openpyxl wheel
import time
import openpyxl
import calendar
import datetime
from openpyxl.styles import Border, Side


# input personal information ##########################################################################################

# for fIle name
print('有个事儿得说一下，每个月的工作日是自动计算出来了，但是并没有将中国法定假日算进去，\n\n所以说，有三天七天假日的那种，希望动动你的小手，老铁双击溜溜溜？？(｡◕ˇ∀ˇ◕)\n')
print('周末下午所在的行存在，不能通过方法删除，需要手动删一些行\n')
print('请严格输入的格式可以参考括号里的提示······\n')
Year  = input('输入要生成日志的年份，如(2020):\n ')
Month = input('输入要生成日志的月份，如(07):\n ')
NameSmp  = input('输入名字缩写，如(aha):\n ')
# for table head
NameCN   = input('输入中文全名，如(狮子王):\n ')
NumWork  = input('输入工号，如(50):\n ')
Department = input('输入所在部门，如(电池管理):\n ')
'''
# tmp info for dbg
print('\n信息临时固定，原谅我懒得实在不想每次dbg时候输入了······\n')
Year  = '2019 '
Month = '11'
NameSmp  = 'dwa'
# for table head
NameCN   = '王柏东'
NumWork  = '50'
Department = '电池管理'
'''
# get the actual date #################################################################################################
NumDat = calendar.monthrange(int(Year),int(Month))[1]	# number of date the month
print('本月有%s天\n月历表' % str(NumDat))

monthcalendar = calendar.monthcalendar(int(Year),int(Month))
MthLst = []
for i in range(0, len(monthcalendar)):	# change to a list
	MthLst += monthcalendar[i]
	print(monthcalendar[i])
Place = monthcalendar[0].index(1)
#print('一号是星期%s' % str(Place+1))

WorkDays = 0
x = 0
for y in range(1, NumDat+1):
	EvryDayPlace = monthcalendar[x].index(y)
	if EvryDayPlace == 0 or EvryDayPlace == 1 or EvryDayPlace == 2 or EvryDayPlace == 3 or EvryDayPlace == 4:
		WorkDays += 1		
	if EvryDayPlace == 6:
		x += 1
print('\n贼个月要上%s天班！(在不考虑法定假日的情况下)\n' % str(WorkDays))	
# creat a Excel file ##################################################################################################
Fil = openpyxl.Workbook()
EngMonth = 'EnMonth'
# make file name
if Month == '01':
	EngMonth = 'Jan'
elif Month == '02':
	EngMonth = 'Feb'
elif Month == '03':
	EngMonth = 'Mar'
elif Month == '04':
	EngMonth = 'Apr'
elif Month == '05':
	EngMonth = 'May'
elif Month == '06':
	EngMonth = 'Jun'
elif Month == '07':
	EngMonth = 'Jul'
elif Month == '08':
	EngMonth = 'Aug'
elif Month == '09':
	EngMonth = 'Sep'
elif Month == '10':
	EngMonth = 'Oct'
elif Month == '11':
	EngMonth = 'Nov'
elif Month == '12':
	EngMonth = 'Dec'	

Filename = 'Working Log %s_%s_%s.xlsx' % (EngMonth, Year[2:4], NameSmp)

# new default sheet
Sht = Fil.active
Sht.title = 'Sheet1'
Sht.freeze_panes = 'A4'	# freeze 1-3 lines

# set format of cell ##################################################################################################
for m in range(1, NumDat*2+3):
	for n in range(0, 8):
		if n == 0:
			Table_Nam_border = 'A%s' % m
		elif n == 1:
			Table_Nam_border = 'B%s' % m
		elif n == 2:
			Table_Nam_border = 'C%s' % m
		elif n == 3:
			Table_Nam_border = 'D%s' % m
		elif n == 4:
			Table_Nam_border = 'E%s' % m
		elif n == 5:
			Table_Nam_border = 'F%s' % m
		elif n == 6:
			Table_Nam_border = 'G%s' % m
		elif n == 7:
			Table_Nam_border = 'H%s' % m	
		#Sht[Table_Nam_border].border = openpyxl.styles.borders.Border(
Sht.column_dimensions['A'].width = 17.37	# +0.62 with true value(only column width needs)
Sht.column_dimensions['B'].width = 9.75
Sht.column_dimensions['C'].width = 9.75
Sht.column_dimensions['D'].width = 9.75
Sht.column_dimensions['E'].width = 9.75
Sht.column_dimensions['F'].width = 9.75
Sht.column_dimensions['G'].width = 24.12
Sht.column_dimensions['H'].width = 97.62
Sht.row_dimensions[1].height = 35
for row_height in range(2,(2*NumDat +17)):
	Sht.row_dimensions[row_height].height = 20

Sht.merge_cells('A1:H1')	# head
Sht['A1'].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
Sht['A1'].font = openpyxl.styles.Font(name='宋体', size=16, bold=True)
ColorFill_Gray = openpyxl.styles.PatternFill(start_color ='C0C0C0', end_color = 'C0C0C0', fill_type = 'solid')
ColorFill_Green = openpyxl.styles.PatternFill(start_color ='00B050', end_color = '00B050', fill_type = 'solid')
ColorFill_Blue = openpyxl.styles.PatternFill(start_color ='00B0F0', end_color = '00B0F0', fill_type = 'solid')
ColorFill_Yellow = openpyxl.styles.PatternFill(start_color ='FFFF00', end_color = 'FFFF00', fill_type = 'solid')

Sht.merge_cells('B2:C2')	# information
Sht.merge_cells('E2:F2')
for MergeCells in range(NumDat*2+6, NumDat*2+9+1):
	MergCells_botm_1 = 'D%s' % MergeCells
	MergCells_botm_2 = 'F%s' % MergeCells
	MergCells_botm_3 = 'G%s' % MergeCells
	MergCells_botm_4 = 'H%s' % MergeCells
	formula_WokHour = '=SUM(G4:G%s)' % (NumDat*2+4)
	formula_WokDays = WorkDays
	formula_NormalWokHours = '=G%s*8' % (NumDat*2+7)
	formula_AddWorkTimHour = '=G%s-G%s' % ((NumDat*2+6), (NumDat*2+8))
	Sht[MergCells_botm_1].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
	Sht[MergCells_botm_1].font      = openpyxl.styles.Font(name='宋体', size=11, color='FF0000', bold=True)	# red color
	Sht[MergCells_botm_3].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
	Sht[MergCells_botm_4].font      = openpyxl.styles.Font(name='宋体', size=11, bold=True)	
	Sht.merge_cells(MergCells_botm_1 + ':' + MergCells_botm_2)
	
	if MergeCells == NumDat*2+6:
		Sht[MergCells_botm_1] = '总净工时(十进制)='
		Sht[MergCells_botm_4] = '小时'
		Sht[MergCells_botm_3].fill = ColorFill_Gray
		Sht[MergCells_botm_3] = formula_WokHour
		Sht[MergCells_botm_3].number_format = '0.000'
	elif MergeCells == NumDat*2+7:
		Sht[MergCells_botm_1] = '工作日='
		Sht[MergCells_botm_4] = '天'
		Sht[MergCells_botm_3] = formula_WokDays
		Sht[MergCells_botm_3].number_format = '0.000'
	elif MergeCells == NumDat*2+8:
		Sht[MergCells_botm_1] = '正常工时(十进制)='
		Sht[MergCells_botm_4] = '小时'
		Sht[MergCells_botm_3].fill = ColorFill_Gray
		Sht[MergCells_botm_3] = formula_NormalWokHours
		Sht[MergCells_botm_3].number_format = '0.000'
	elif MergeCells == NumDat*2+9:
		Sht[MergCells_botm_1] = '加班时间(十进制)='
		Sht[MergCells_botm_4] = '小时'
		Sht[MergCells_botm_3].fill = ColorFill_Gray
		Sht[MergCells_botm_3] = formula_AddWorkTimHour
		Sht[MergCells_botm_3].number_format = '0.000'

signature = 'G%s' % (NumDat*2+11)
signaturedate = 'G%s' % (NumDat*2+12)
Sht[signature] = '主管签字：'
Sht[signaturedate] = '签字日期：'
Sht[signature].font = openpyxl.styles.Font(name='宋体', size=11, bold=True)
Sht[signature].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
Sht[signaturedate].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")
Sht[signaturedate].font = openpyxl.styles.Font(name='宋体', size=11, bold=True)


# Note
Note_Merge_Beg = 'D%s' % (NumDat*2+14)
Note_Merge_End = 'D%s' % (NumDat*2+16)
Note_ColorGreen = 'E%s' % (NumDat*2+14)
Note_ColorYellow = 'E%s' % (NumDat*2+15)
Note_ColorBlue = 'E%s' % (NumDat*2+16)
Note_Merge_Right_1_Beg = 'F%s' % (NumDat*2+14)
Note_Merge_Right_1_End = 'G%s' % (NumDat*2+14)
Note_Merge_Right_2_Beg = 'F%s' % (NumDat*2+15)
Note_Merge_Right_2_End = 'G%s' % (NumDat*2+15)
Note_Merge_Right_3_Beg = 'F%s' % (NumDat*2+16)
Note_Merge_Right_3_End = 'G%s' % (NumDat*2+16)
Sht.merge_cells(Note_Merge_Beg + ':' + Note_Merge_End)
Sht.merge_cells(Note_Merge_Right_1_Beg + ':' + Note_Merge_Right_1_End)
Sht.merge_cells(Note_Merge_Right_2_Beg + ':' + Note_Merge_Right_2_End)
Sht.merge_cells(Note_Merge_Right_3_Beg + ':' + Note_Merge_Right_3_End)
Sht[Note_Merge_Beg] = '注释：'
Sht[Note_Merge_Right_1_Beg] = '休假（带薪假）'
Sht[Note_Merge_Right_2_Beg] = '周末 & 法定假日'
Sht[Note_Merge_Right_3_Beg] = '病假'
Sht[Note_ColorGreen].fill = ColorFill_Green
Sht[Note_ColorYellow].fill = ColorFill_Yellow
Sht[Note_ColorBlue].fill = ColorFill_Blue


Sht[Note_Merge_Beg].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
Sht[Note_Merge_Beg].font      = openpyxl.styles.Font(name='宋体', size=11, color='FF0000', bold=True)
Sht[Note_Merge_Right_1_Beg].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
Sht[Note_Merge_Right_2_Beg].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
Sht[Note_Merge_Right_3_Beg].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

x_Yellow = 0
for FillColorsYellow in range(1, NumDat+1):	# Weekend in Yellow fill
	EvryDayPlace_Yellow = monthcalendar[x_Yellow].index(FillColorsYellow)	
	if EvryDayPlace_Yellow == 5 or EvryDayPlace_Yellow == 6:
		YellowDay_1 = 'A%s' % (FillColorsYellow*2+2)
		YellowDay_2 = 'B%s' % (FillColorsYellow*2+2)
		YellowDay_3 = 'C%s' % (FillColorsYellow*2+2)
		YellowDay_4 = 'D%s' % (FillColorsYellow*2+2)
		YellowDay_5 = 'E%s' % (FillColorsYellow*2+2)
		YellowDay_6 = 'F%s' % (FillColorsYellow*2+2)
		YellowDay_7 = 'G%s' % (FillColorsYellow*2+2)
		YellowDay_8 = 'H%s' % (FillColorsYellow*2+2)
		Sht[YellowDay_1].fill = ColorFill_Yellow
		Sht[YellowDay_2].fill = ColorFill_Yellow
		Sht[YellowDay_3].fill = ColorFill_Yellow
		Sht[YellowDay_4].fill = ColorFill_Yellow
		Sht[YellowDay_5].fill = ColorFill_Yellow
		Sht[YellowDay_6].fill = ColorFill_Yellow
		Sht[YellowDay_7].fill = ColorFill_Yellow
		Sht[YellowDay_8].fill = ColorFill_Yellow		
	if EvryDayPlace_Yellow == 6:
		x_Yellow += 1

		
		
for row in range(2,4):
	for column in range(0,8):
		if column == 0:
			Table_Nam = 'A%s' % row
		elif column == 1:
			Table_Nam = 'B%s' % row
		elif column == 2:
			Table_Nam = 'C%s' % row
		elif column == 3:
			Table_Nam = 'D%s' % row
		elif column == 4:
			Table_Nam = 'E%s' % row
		elif column == 5:
			Table_Nam = 'F%s' % row
		elif column == 6:
			Table_Nam = 'G%s' % row
		elif column == 7:
			Table_Nam = 'H%s' % row	
			
		#print(Table_Nam)
		Sht[Table_Nam].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
		Sht[Table_Nam].font      = openpyxl.styles.Font(name='宋体', size=11, bold=True)

# column 'Date & weekday'
for LopMth in range(0, NumDat):
	Date = MthLst[Place + LopMth]
	col_nam_A = 'A%s' % (Date*2+2)
	Sht[col_nam_A].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")	# calendar format
	#Sht[col_nam_A].number_format = 'yyyy-mm-dd'
	Sht[col_nam_A].number_format = 'yyyy"年"mm"月"dd"日"'
	
	col_nam_B = 'B%s' % (Date*2+2)
	Sht[col_nam_B].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

# column 'Times * 3'


# fill the table ######################################################################################################
#for raw in range(1,80):	# loop to fill table
#	Sht.append(range(10))

Sht['A1'] = '希科泰%s年%s月工作日志' % (Year, Month)	# table head
Sht['A1'].number_format = '希科泰0000年00月工作日志'	# format
Sht['A2'] = '姓名：'
Sht['B2'] = '%s' % NameCN
Sht['D2'] = '工号：'
if (int(NumWork)) < 100:
	Sht['E2'] = '0000%s' % NumWork
elif (int(NumWork)) > 99 and (int(NumWork)) < 1000:
	Sht['E2'] = '000%s' % NumWork
Sht['G2'] = '部门：'
Sht['H2'] = '%s' % Department
Sht['A3'] = '日期'
Sht['B3'] = '星期'
Sht['C3'] = '开始时间'
Sht['D3'] = '结束时间'
Sht['E3'] = '休息时间'
Sht['F3'] = '净工时'
Sht['G3'] = '净工时（十进制）/小时'
Sht['H3'] = '工作内容'

# column 'Date & weekday'
WekDat = Place
for LopMth in range(0, NumDat):
	Date = MthLst[Place + LopMth]
	col_nam_A = 'A%s' % (Date*2+2)
	Sht[col_nam_A] = datetime.datetime(int(Year), int(Month), int(Date))	# calendar	
	col_nam_B = 'B%s' % (Date*2+2)
	if WekDat == 0:
		Sht[col_nam_B] = '星期一'	# week date
	elif WekDat == 1:
		Sht[col_nam_B] = '星期二'
	elif WekDat == 2:
		Sht[col_nam_B] = '星期三'
	elif WekDat == 3:
		Sht[col_nam_B] = '星期四'
	elif WekDat == 4:
		Sht[col_nam_B] = '星期五'
	elif WekDat == 5:
		Sht[col_nam_B] = '星期六'
	elif WekDat == 6:
		Sht[col_nam_B] = '星期日'
	WekDat += 1
	if WekDat > 6:
		WekDat = 0

for WrkTim in range(4, NumDat*2+4):
	col_WrkTim_C = 'C%s' % WrkTim
	col_WrkTim_D = 'D%s' % WrkTim
	col_WrkTim_E = 'E%s' % WrkTim
	col_WrkTim_F = 'F%s' % WrkTim
	col_WrkTim_G = 'G%s' % WrkTim
	Sht[col_WrkTim_C] = '00:00'
	Sht[col_WrkTim_D] = '00:00'
	Sht[col_WrkTim_E] = '00:00'
	Sht[col_WrkTim_F] = '=D%s-C%s-E%s' % (WrkTim, WrkTim, WrkTim)
	Sht[col_WrkTim_G] = '=HOUR(F%s)+MINUTE(F%s)/60' % (WrkTim, WrkTim)
	Sht[col_WrkTim_C].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")	# calendar format
	Sht[col_WrkTim_D].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")	# calendar format
	Sht[col_WrkTim_E].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")	# calendar format
	Sht[col_WrkTim_F].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")	# calendar format
	Sht[col_WrkTim_G].alignment = openpyxl.styles.Alignment(horizontal="right", vertical="center")	# calendar format
	Sht[col_WrkTim_C].number_format = 'hh:mm'
	Sht[col_WrkTim_D].number_format = 'hh:mm'
	Sht[col_WrkTim_E].number_format = 'hh:mm'
	Sht[col_WrkTim_F].number_format = 'hh:mm'
	Sht[col_WrkTim_G].number_format = '0.000'

	
# improve format ######################################################################################################
bian = Side(style='thin', color='000000')
border = Border(top=bian, bottom=bian, left=bian, right=bian)
for BorderLine in range(1, NumDat*2+4):
	TabBorNam_A = 'A%s' % BorderLine
	TabBorNam_B = 'B%s' % BorderLine
	TabBorNam_C = 'C%s' % BorderLine
	TabBorNam_D = 'D%s' % BorderLine
	TabBorNam_E = 'E%s' % BorderLine
	TabBorNam_F = 'F%s' % BorderLine
	TabBorNam_G = 'G%s' % BorderLine
	TabBorNam_H = 'H%s' % BorderLine
	Sht[TabBorNam_A].border=border
	Sht[TabBorNam_B].border=border
	Sht[TabBorNam_C].border=border
	Sht[TabBorNam_D].border=border
	Sht[TabBorNam_E].border=border
	Sht[TabBorNam_F].border=border
	Sht[TabBorNam_G].border=border
	Sht[TabBorNam_H].border=border

for BorderWorkTim in range(NumDat*2+6, NumDat*2+10):
	TabBorNam_D = 'D%s' % BorderWorkTim
	TabBorNam_E = 'E%s' % BorderWorkTim
	TabBorNam_F = 'F%s' % BorderWorkTim
	TabBorNam_G = 'G%s' % BorderWorkTim
	Sht[TabBorNam_D].border=border
	Sht[TabBorNam_E].border=border
	Sht[TabBorNam_F].border=border
	Sht[TabBorNam_G].border=border

for BorderNote in range(NumDat*2+14, NumDat*2+17):
	TabBorNam_D = 'D%s' % BorderNote
	TabBorNam_E = 'E%s' % BorderNote
	TabBorNam_F = 'F%s' % BorderNote
	TabBorNam_G = 'G%s' % BorderNote
	Sht[TabBorNam_D].border=border
	Sht[TabBorNam_E].border=border
	Sht[TabBorNam_F].border=border
	Sht[TabBorNam_G].border=border

	

# save file ###########################################################################################################
Fil.save(filename = Filename)
'''
LodBok = openpyxl.load_workbook(filename = Filename)
LodSht = LodBok['Sheet1']

for LopLodSht in range(4, NumDat*2+4):
	NxtLine = LopLodSht + 1
	LodTabNam = 'B%s' % LopLodSht
	WeekDayNum = LodSht[LodTabNam].value
	if WeekDayNum == '星期六' or WeekDayNum == '星期日':
		Sht.delete_rows(NxtLine, 1)

Fil.save(filename = Filename)
'''
print('Yes！%s年%s月份工作日志模板生成完毕！' % (Year, Month))
print('5秒后自动退出，该干嘛干嘛去吧······')
time.sleep(5)